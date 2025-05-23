import openpyxl
import csv

file_path = "IL.xlsx" 
wb = openpyxl.load_workbook(file_path)
ws = wb.active

for merged_range in list(ws.merged_cells.ranges):
    merged_cell = merged_range.start_cell
    value = merged_cell.value
    ws.unmerge_cells(str(merged_range))
    for row_col_indices in merged_range.cells:
        ws.cell(*row_col_indices).value = value

# Converging abbreviations to full names
conversion_wb = openpyxl.load_workbook("dosage_form_conversion.xlsx", data_only=True)
conversion_ws = conversion_wb.active
conversion_dict = {}
for row in conversion_ws.iter_rows(min_row=2, values_only=True):
    abbr, full = row
    if abbr and full:
        conversion_dict[str(abbr).strip()] = str(full).strip()

header_keywords = {"Drug Class", "Drug Name", "Dosage Form", "PDL Status"}
header_row = None
column_indices = {"PDL Status": []}

for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if any(cell in header_keywords for cell in row if cell):
        header_row = idx
        for col_idx, cell in enumerate(row, start=1):
            if cell == "PDL Status":
                column_indices["PDL Status"].append(col_idx)
            elif cell in {"Drug Class", "Drug Name", "Dosage Form"}:
                column_indices[cell] = col_idx
        break

if not header_row or not column_indices["PDL Status"]:
    raise ValueError("Header row not found or required columns are missing.")

output = [["therapeutic_class", "pdl_name", "status"]]
last_therapeutic_class = None

def get_cell(row, col_name):
    idx = column_indices.get(col_name)
    return str(row[idx - 1]).strip() if idx and row[idx - 1] else ""

for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    therapeutic_class = get_cell(row, "Drug Class") or last_therapeutic_class
    if therapeutic_class:
        last_therapeutic_class = therapeutic_class
    else:
        continue

    drug_name = get_cell(row, "Drug Name")
    dosage_abbr = get_cell(row, "Dosage Form")
    dosage_full = conversion_dict.get(dosage_abbr, dosage_abbr)

    status = ""
    for col_idx in column_indices["PDL Status"]:
        value = row[col_idx - 1]
        if isinstance(value, str) and value.strip():
            normalized = value.strip().lower()
            if normalized in {"preferred", "preferred_with_pa"}:
                status = "Preferred"
            elif normalized == "non_preferred":
                status = "Non-Preferred"
            else:
                status = value.strip()
            break

    pdl_name = f"{drug_name} {dosage_full}".strip()

    output.append([
        therapeutic_class if therapeutic_class else "Error",
        pdl_name if pdl_name else "Error",
        status if status else "Error"
    ])

with open("IL_PDL.csv", "w", newline='', encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerows(output)