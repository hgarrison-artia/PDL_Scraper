import openpyxl
import csv
import re

file_path = "ME.xlsx" 
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Unmerge all merged cells
for merged_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged_range))

# Remove superscripts from cell texts
SUPERSCRIPT_UNICODE = "⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾"
def remove_superscripts(text):
    if text and isinstance(text, str):
        text = re.sub(f"[\u2070-\u209F]", "", text)
        return " ".join(text.split())
    return text

# Define the Excel indexed colors because openpyxl does not provide a direct way to get the color majority of the time
# We use colors to recognize "headers"; the headers have been mint-colored rows until now (April 2025)
EXCEL_INDEXED_COLORS = {
    0: "000000", 1: "FFFFFF", 2: "FF0000", 3: "00FF00", 4: "0000FF", 
    5: "FFFF00", 6: "FF00FF", 7: "00FFFF", 8: "000000", 9: "FFFFFF",
    10: "800000", 11: "008000", 12: "000080", 13: "808000", 14: "800080",
    15: "008080", 16: "C0C0C0", 17: "808080", 18: "9999FF", 19: "993366",
    20: "FFFFCC", 21: "CCFFFF", 22: "660066", 23: "FF8080", 24: "0066CC",
    25: "CCCCFF", 26: "000080", 27: "FF00FF", 28: "FFFF00", 29: "00FFFF",
    30: "800080", 31: "800000", 32: "008080", 33: "0000FF", 34: "00CCFF",
    35: "CCFFFF", 36: "CCFFCC", 37: "FFFF99", 38: "99CCFF", 39: "FF99CC",
    40: "CC99FF", 41: "FFCC99", 42: "3366FF", 43: "33CCCC", 44: "99CC00",
    45: "FFCC00", 46: "FF9900", 47: "FF6600", 48: "666699", 49: "969696"
}

header_rows = set()

# Iterate through all rows in column 1
for row in ws.iter_rows(min_col=1, max_col=1):
    for cell in row:
        fgColor = cell.fill.fgColor
        fill_color = "No Fill"
        if fgColor is not None:
            if hasattr(fgColor, "rgb") and isinstance(fgColor.rgb, str):
                fill_color = fgColor.rgb  # Direct RGB
            elif hasattr(fgColor, "indexed") and fgColor.indexed is not None:
                indexed_value = fgColor.indexed
                fill_color = EXCEL_INDEXED_COLORS.get(indexed_value, f"Unknown Indexed Color {indexed_value}")

        # Using RBG code 969696 for mint color but it needs revision if there is any change in row colors
        if fill_color.lower().endswith("969696"):
            header_rows.add(cell.row)

header_rows = sorted(header_rows)    # Identifying the header rows
start_row = header_rows[0]           # Starting from the first header row

column_headers = ["Therapeutic Category", "Preferred", "Non-Preferred"]
data = [column_headers]
column_indices = {}

for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
    for col_idx, cell in enumerate(row, start=1):
        column_indices[cell] = col_idx
    break

def validate_text(value):
    return isinstance(value, str) and bool(value.strip())  # Ensure it's a non-empty string

last_category = None

# Extract data from the relevant columns based on dynamic header detection
for row_idx, row in enumerate(ws.iter_rows(min_row=start_row + 1, values_only=True), start=start_row + 1):
    if row_idx in header_rows:
        continue  # Skip header rows

    drug_class_idx = column_indices.get("CATEGORY")
    preferred_idx = column_indices.get("PREFERRED DRUGS")
    non_preferred_idx = column_indices.get("NON-PREFERRED DRUGS       PA Required")

    therapeutic_class = remove_superscripts(row[drug_class_idx - 1]) if row[drug_class_idx - 1] else None
    if therapeutic_class:
        last_category = therapeutic_class
    else:
        therapeutic_class = last_category

    preferred = remove_superscripts(row[preferred_idx - 1]) if preferred_idx else ''
    non_preferred = remove_superscripts(row[non_preferred_idx - 1]) if non_preferred_idx else ''

    if not preferred and not non_preferred:
        continue
    if isinstance(preferred, str) and preferred.isupper() and isinstance(non_preferred, str) and non_preferred.isupper() and preferred==non_preferred:
        continue

    extracted_row = [
        remove_superscripts(therapeutic_class),
        preferred, 
        non_preferred
    ]

    extracted_row = [value if validate_text(value) else "" for value in extracted_row]
    data.append(extracted_row)

for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            cell.value = remove_superscripts(cell.value)

# Export data to CSV
csv_file_path = "../ME_PDL.csv"
with open(csv_file_path, mode="w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)
    writer.writerows(data)

print(f"Data successfully exported to {csv_file_path}")


