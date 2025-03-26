import openpyxl
import csv

file_path = "FL.xlsx" 
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Unmerge all merged cells
for merged_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged_range))

# Keywords to identify the header row
header_keywords = {"HIC3", "HIC3 Desc", "Label Name", "Generic Name", 
                   "Medicaid Min Age", "Medicaid Max Age", "PA Required"}

# Find the row index where the headers appear
header_row = None
column_indices = {}

# Find the header row and identify column indices for the relevant headers
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
    if row and any(cell in header_keywords for cell in row if cell):  # Ignore empty cells
        header_row = idx
        # Identify the column index of each required header
        for col_idx, cell in enumerate(row, start=1):
            if cell in header_keywords:
                column_indices[cell] = col_idx
        break

if not header_row or not column_indices:
    raise ValueError("Header row not found or required columns are missing. Please check the file structure.")

# Define the output CSV column headers
column_headers = ["Therapeutic Category", "Label Name", "Generic Name", "Min Age", "Max Age", "PA Required"]

# Initialize data list
data = [column_headers]
last_category = None  # To track the last seen Therapeutic Category

# Validation functions
def validate_text(value):
    return isinstance(value, str) and bool(value.strip())  # Ensure it's a non-empty string

def validate_numeric(value):
    try:
        float(value)  # Check if it's numeric (int or float)
        return True
    except ValueError:
        return False

def validate_yes_no(value):
    return value in ["Yes", "No"]

# Extract data from the relevant columns based on dynamic header detection
for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    # Extract the relevant columns based on the header
    extracted_row = [
        row[column_indices.get("HIC3 Desc", None) - 1] if "HIC3 Desc" in column_indices else None,
        row[column_indices.get("Label Name", None) - 1] if "Label Name" in column_indices else None,
        row[column_indices.get("Generic Name", None) - 1] if "Generic Name" in column_indices else None,
        row[column_indices.get("Medicaid Min Age", None) - 1] if "Medicaid Min Age" in column_indices else None,
        row[column_indices.get("Medicaid Max Age", None) - 1] if "Medicaid Max Age" in column_indices else None,
        row[column_indices.get("PA Required", None) - 1] if "PA Required" in column_indices else None,
    ]
    
    # Validate each column before assigning
    if extracted_row[0] and not validate_text(extracted_row[0]):
        extracted_row[0] = "Error"  # Set invalid values to "Error"
    if extracted_row[1] and not validate_text(extracted_row[1]):
        extracted_row[1] = "Error"  # Set invalid values to "Error"
    if extracted_row[2] and not validate_text(extracted_row[2]):
        extracted_row[2] = "Error"  # Set invalid values to "Error"
    if extracted_row[3] and not validate_numeric(extracted_row[3]):
        extracted_row[3] = "Error"  # Set invalid values to "Error"
    if extracted_row[4] and not validate_numeric(extracted_row[4]):
        extracted_row[4] = "Error"  # Set invalid values to "Error"
    if extracted_row[5] and not validate_text(extracted_row[5]):
        extracted_row[5] = "Error"  # Set invalid values to "Error"

    # Ensure Therapeutic Category is filled down
    if extracted_row[0]:  # If there's a category, update last_category
        last_category = extracted_row[0]
    else:
        extracted_row[0] = last_category  # Fill with the last seen category

    data.append(extracted_row)

# Export data to CSV
csv_file_path = "FL_PDL.csv" 
with open(csv_file_path, mode="w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)
    writer.writerows(data)

print(f"Data successfully exported to {csv_file_path}")
