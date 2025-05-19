import pandas as pd
from openpyxl import load_workbook

file_path = "CT.xlsx"
wb = load_workbook(filename=file_path, data_only=True)
ws = wb.active

# RGB value for grey fill that identifies therapeutic classes
THERAPEUTIC_GREY = 'FFBEBEBE'

status = "Preferred"  # Default status
records = []

for col in range(1, ws.max_column + 1):
    current_class = None
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        fill = cell.fill

        if not value:
            continue

        # Check for grey therapeutic class cell
        if fill and fill.start_color and fill.start_color.rgb == THERAPEUTIC_GREY:
            current_class = value
        elif current_class:
            records.append((current_class, value, status))

df = pd.DataFrame(records, columns=['therapeutic_class', 'pdl_name', 'status'])

# Filter out rows where any column contains text surrounded by asterisks
mask = ~df.apply(lambda x: x.astype(str).str.contains('\*\*\*.*\*\*\*').any(), axis=1)
df = df[mask]

df.to_csv("CT_PDL.csv", index=False)