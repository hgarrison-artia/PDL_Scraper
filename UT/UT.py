import pandas as pd
import openpyxl


wb = openpyxl.load_workbook("UT.xlsx")

Class_Color = 'FFFFD966'
SubClass_Color = 'FFD9D9D9'
drugs = []

previous_class = None
previous_subclass = None

for sheet_name in wb.sheetnames[1:]:
    ws = wb[sheet_name]
    current_class = previous_class
    current_subclass = previous_subclass

    for row in ws.iter_rows():
        # Update class/subclass if found in this row, otherwise keep previous
        for cell in row:
            value = cell.value
            fill = cell.fill
            if fill and fill.fill_type == 'solid' and fill.start_color.rgb == Class_Color:
                current_class = value
                previous_class = value
            elif fill and fill.fill_type == 'solid' and fill.start_color.rgb == SubClass_Color:
                current_subclass = value
                previous_subclass = value

        # Extract drug name from first cell if not class/subclass color
        drug = None
        drugcell = row[0]
        value = drugcell.value
        fill = drugcell.fill
        if value and not (
            fill and fill.fill_type == 'solid' and (
                fill.start_color.rgb == Class_Color or fill.start_color.rgb == SubClass_Color
            )
        ):
            drug = value

        # Extract status from second cell if not class/subclass color
        status = None
        if len(row) > 1:
            statuscell = row[1]
            status_value = statuscell.value
            status_fill = statuscell.fill
            if (
                status_value
                and str(status_value).strip() in ["Preferred", "Non Preferred"]
                and not (
                    status_fill and status_fill.fill_type == 'solid' and (
                        status_fill.start_color.rgb == Class_Color or status_fill.start_color.rgb == SubClass_Color
                    )
                )
            ):
                status = status_value

        # Only append if drug and status are found
        if drug and status:
            drugs.append((current_class, current_subclass, drug, status))
            


df = pd.DataFrame(drugs, columns=['therapeutic_class', 'Subclass', 'pdl_name', 'status'])
df['therapeutic_class'] = df['therapeutic_class'].fillna('').astype(str) + ': ' + df['Subclass'].fillna('').astype(str)
df = df.drop(columns=['Subclass'])

df['status'] = ["Non-Preferred" if status == "Non Preferred" else "Preferred" for status in df['status']]
df['pdl_name'] = [name.replace("\n", " ") for name in df['pdl_name']]
df.to_csv('UT_PDL.csv', index=False)